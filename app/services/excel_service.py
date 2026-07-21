"""
Сервис импорта и экспорта суточных рапортов в формате Excel (.xlsx).

Импорт: массовая загрузка рапортов из файла оператора (шахматки).
Экспорт: сводный отчёт по скважинам за выбранный месяц для руководства.
Используется openpyxl — стандартная библиотека Python для работы с .xlsx.
"""

import calendar
from datetime import date
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models import DailyProduction, OilCompany, Well
from app.schemas import DailyProductionCreate

# Формула чистой нефти на уровне SQL (как в дашборде)
OIL_EXPR = (
    DailyProduction.liquid_volume
    * (1 - DailyProduction.water_cut / 100)
    * DailyProduction.density
)

# Ожидаемые колонки файла импорта (порядок важен)
IMPORT_COLUMNS = [
    "well_name",
    "date",
    "working_hours",
    "liquid_volume",
    "water_cut",
    "density",
]


def build_import_template(wells=None) -> bytes:
    """
    Создаёт шаблон Excel для импорта с заголовками и примером строки.

    Если передан список скважин, добавляет второй лист «Скважины» —
    справочник названий, чтобы оператор знал что вписывать в колонку Скважина.

    Args:
        wells (list[Well] | None): Список скважин для справочного листа.

    Returns:
        bytes: Содержимое .xlsx файла-шаблона.
    """
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(color="FFFFFF", bold=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Рапорты"

    headers = [
        "Скважина",
        "date (ГГГГ-ММ-ДД)",
        "working_hours (0-24)",
        "liquid_volume",
        "water_cut (0-100)",
        "density (0.7-1.0)",
    ]
    ws.append(headers)

    # Пример строки с реальным названием первой скважины, если она есть
    example_name = wells[0].name if wells else "Скважина-1"
    ws.append([example_name, "2026-06-15", 23.5, 150.0, 25.0, 0.86])

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    # Справочный лист со скважинами и их ID
    if wells:
        ref = wb.create_sheet("Скважины")
        ref.append(["Название", "Тип", "Компания"])
        for w in wells:
            company = w.oil_company.name if w.oil_company else "—"
            ref.append([w.name, w.type, company])
        for cell in ref[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        widths = [20, 16, 30]
        for i, width in enumerate(widths, start=1):
            ref.column_dimensions[chr(64 + i)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _parse_date(value):
    """Превращает значение ячейки в date (поддерживает datetime и строку)."""
    if isinstance(value, date):
        return value
    if hasattr(value, "date"):  # datetime
        return value.date()
    return date.fromisoformat(str(value).strip())


def import_productions(
    db: Session, file_bytes: bytes, company_id: int | None = None
) -> dict:
    """
    Массовый импорт рапортов из Excel-файла.

    Каждая строка валидируется через схему и проверяется на дубликат.
    Некорректные строки пропускаются с указанием причины.

    Args:
        db (Session): Сессия базы данных.
        file_bytes (bytes): Содержимое загруженного .xlsx файла.
        company_id (int | None): Если задан — разрешены только скважины
            этой компании (manager/operator не могут импортировать чужие).

    Returns:
        dict: {"created": int, "updated": int, "errors": [строки с ошибками]}
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    created = 0
    updated = 0
    errors = []

    # row_index начинается с 2 (1-я строка — заголовки)
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # пропускаем полностью пустые строки
        if row is None or all(cell is None for cell in row):
            continue

        # поиск скважины по названию (напр. "Скважина-13AI")
        well_name = str(row[0]).strip() if row[0] is not None else ""
        well = db.query(Well).filter(Well.name == well_name).first()
        if not well:
            errors.append(f"Строка {row_index}: скважина '{well_name}' не найдена")
            continue

        # запрет импорта по чужим скважинам (для manager и operator)
        if company_id is not None and well.oil_company_id != company_id:
            errors.append(
                f"Строка {row_index}: скважина '{well_name}' "
                f"не принадлежит вашей компании"
            )
            continue

        try:
            _, raw_date, hours, liquid, water, density = row[:6]

            # плотность по умолчанию, если ячейка пустая
            if density is None:
                density = 0.86

            data = DailyProductionCreate(
                well_id=well.id,
                date=_parse_date(raw_date),
                working_hours=float(hours),
                liquid_volume=float(liquid),
                water_cut=float(water),
                density=float(density),
            )
        except Exception as e:
            errors.append(f"Строка {row_index}: ошибка данных ({e})")
            continue

        # upsert: если рапорт за эту скважину+дату уже есть — обновляем,
        # иначе создаём новый
        existing = (
            db.query(DailyProduction)
            .filter(
                DailyProduction.well_id == data.well_id,
                DailyProduction.date == data.date,
            )
            .first()
        )

        if existing:
            existing.working_hours = data.working_hours
            existing.liquid_volume = data.liquid_volume
            existing.water_cut = data.water_cut
            existing.density = data.density
            updated += 1
        else:
            report = DailyProduction(
                well_id=data.well_id,
                date=data.date,
                working_hours=data.working_hours,
                liquid_volume=data.liquid_volume,
                water_cut=data.water_cut,
                density=data.density,
            )
            db.add(report)
            created += 1

    db.commit()
    return {"created": created, "updated": updated, "errors": errors}


def export_monthly_summary(
    db: Session, year: int, month: int, company_id: int | None = None
) -> bytes:
    """
    Формирует сводный отчёт по скважинам за месяц в .xlsx.

    Для каждой скважины с рапортами в указанном месяце считает:
    количество рапортов, суммарную жидкость, среднюю обводненность,
    суммарную чистую нефть.

    Args:
        db (Session): Сессия базы данных.
        year (int): Год.
        month (int): Месяц (1-12).

    Returns:
        bytes: Содержимое .xlsx файла отчёта.
    """
    start = date(year, month, 1)
    end = date(year, month, calendar.monthrange(year, month)[1])

    rows = (
        db.query(
            Well.name,
            OilCompany.name,
            Well.type,
            func.count(DailyProduction.id),
            func.sum(DailyProduction.liquid_volume),
            func.avg(DailyProduction.water_cut),
            func.sum(OIL_EXPR),
        )
        .join(DailyProduction, DailyProduction.well_id == Well.id)
        .join(OilCompany, OilCompany.id == Well.oil_company_id)
        .filter(DailyProduction.date >= start, DailyProduction.date <= end)
    )
    if company_id is not None:
        rows = rows.filter(Well.oil_company_id == company_id)
    rows = (
        rows.group_by(Well.name, OilCompany.name, Well.type)
        .order_by(func.sum(OIL_EXPR).desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"Сводка {month:02d}.{year}"

    title = f"Сводный отчёт по скважинам за {month:02d}.{year}"
    ws.append([title])
    ws.merge_cells("A1:G1")
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])

    headers = [
        "Скважина",
        "Компания",
        "Тип",
        "Кол-во рапортов",
        "Сумма жидкости (т)",
        "Средняя обводненность (%)",
        "Сумма нефти (т)",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append(
            [
                r[0],
                r[1],
                r[2],
                r[3],
                round(r[4] or 0, 2),
                round(r[5] or 0, 1),
                round(r[6] or 0, 2),
            ]
        )

    # авто-ширина колонок
    widths = [18, 28, 16, 16, 20, 26, 16]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_detailed_reports(
    db: Session, year: int, month: int, company_id: int | None = None
) -> bytes:
    """
    Детальная выгрузка рапортов за месяц — каждый рапорт отдельной строкой.

    Колонки совпадают с шаблоном импорта (Скважина, date, working_hours,
    liquid_volume, water_cut, density), поэтому выгруженный файл можно
    отредактировать и загрузить обратно через импорт (режим upsert).

    Args:
        db (Session): Сессия базы данных.
        year (int): Год.
        month (int): Месяц (1-12).

    Returns:
        bytes: Содержимое .xlsx файла.
    """
    start = date(year, month, 1)
    end = date(year, month, calendar.monthrange(year, month)[1])

    rows = (
        db.query(
            Well.name,
            DailyProduction.date,
            DailyProduction.working_hours,
            DailyProduction.liquid_volume,
            DailyProduction.water_cut,
            DailyProduction.density,
        )
        .join(Well, Well.id == DailyProduction.well_id)
        .filter(DailyProduction.date >= start, DailyProduction.date <= end)
    )
    if company_id is not None:
        rows = rows.filter(Well.oil_company_id == company_id)
    rows = rows.order_by(DailyProduction.date.desc(), DailyProduction.id.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Рапорты"

    # ровно те же заголовки, что у шаблона импорта
    headers = [
        "Скважина",
        "date (ГГГГ-ММ-ДД)",
        "working_hours (0-24)",
        "liquid_volume",
        "water_cut (0-100)",
        "density (0.7-1.0)",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append(
            [
                r[0],
                r[1].isoformat(),
                r[2],
                r[3],
                r[4],
                r[5],
            ]
        )

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
